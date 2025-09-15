import json
from odoo import http
from odoo.http import request
from datetime import datetime
import xlsxwriter
import io
import openpyxl
from openpyxl import load_workbook
import logging
_logger = logging.getLogger(__name__)

class EmployeeAPI(http.Controller):

    @http.route('/api/hr/employees', type='http', auth='user', methods=['POST'], csrf=False)
    def create_employee(self, **kwargs):
        try:
            if request.httprequest.data:
                data = json.loads(request.httprequest.data.decode("utf-8"), strict=False)
            else:
                data = kwargs 

            required_fields = ["name", "birthday", "gender", "work_phone", "work_email", "department_id", "job_id", "id_number", "id_issued_place", "id_issued_date", "permanent_address"]
            for field in required_fields:
                if field not in data or not data[field]:
                    return request.make_response(
                        json.dumps({"error": f"Vui lòng nhập field: {field}"}),
                        headers=[('Content-Type', 'application/json')],
                        status=400
                    )

            employee = request.env["hr.employee"].sudo().create({
                "code": data.get("code"),
                "name": data.get("name"),
                "birthday": data.get("birthday"),
                "gender": data.get("gender"),
                "work_phone": data.get("work_phone"),
                "work_email": data.get("work_email"),
                "department_id": data.get("department_id"),
                "job_id": data.get("job_id"),
                "id_number": data.get("id_number"),
                "id_issued_place": data.get("id_issued_place"),
                "id_issued_date": data.get("id_issued_date"),
                "permanent_address": data.get("permanent_address"),
                "temporary_address": data.get("temporary_address"),
                "tax_id": data.get("tax_id"),
                "insurance_id": data.get("insurance_id"),
                "bank_account": data.get("bank_account"),
            })

            contract_data = data.get("contract")[0] if data.get("contract") and isinstance(data.get("contract"), list) and len(data.get("contract")) > 0 else None
            if contract_data:
                request.env["hr.contract.custom"].sudo().create({
                    "name": contract_data.get("name"),
                    "contract_type": contract_data.get("contract_type"),
                    "contract_term": contract_data.get("contract_term"),
                    "employee_id": employee.id,
                    "date_start": contract_data.get("date_start"),
                    "date_end": contract_data.get("date_end"),
                    "wage": contract_data.get("wage"),
                    "bonus": contract_data.get("bonus")
                })

            return request.make_response(
                json.dumps({
                    "message": "Tạo nhân viên thành công",
                }, ensure_ascii=False),
                headers=[('Content-Type', 'application/json')]
            )
        except Exception as e:
            # Rollback transaction in case of error
            request.env.cr.rollback()

            return request.make_response(
                json.dumps({"error": f"Tạo nhân viên thất bại: {str(e)}"}, ensure_ascii=False),
                headers=[('Content-Type', 'application/json')],
                status=500
            )

    @http.route('/api/hr/employees', type='http', auth='user', methods=['GET'], csrf=False)
    def list_employees(self, **kwargs):
        q = kwargs.get('q', '')
        department_id = kwargs.get('department_id')
        job_id = kwargs.get('job_id')
        gender = kwargs.get('gender', '')
        contract_type = kwargs.get('contract_type', '')
        page = int(kwargs.get('page', 1))
        limit = int(kwargs.get('limit', 25))

        domain = []

        if q:
            domain.append(('name', 'ilike', q))

        if department_id:
            domain.append(('department_id', '=', int(department_id)))

        if job_id:
            domain.append(('job_id', '=', int(job_id)))

        if gender:
            domain.append(('gender', '=', gender))
        
        if contract_type:
            employee_ids = request.env['hr.contract.custom'].sudo().search([('contract_type', '=', contract_type)]).mapped('employee_id').ids
            domain.append(('id', 'in', employee_ids))

        total = request.env['hr.employee'].sudo().search_count(domain)

        offset = (page - 1) * limit

        employees = request.env['hr.employee'].sudo().search(domain, offset=offset, limit=limit, order='code asc')

        data = []
        for emp in employees:
            data.append({
                "id": emp.id,
                "code": emp.code or None,
                "name": emp.name or None,
                "birthday": str(emp.birthday) if emp.birthday else None,
                "gender": emp.gender or None,
                "work_phone": emp.work_phone or None,
                "work_email": emp.work_email or None,
                "department_id": emp.department_id.id if emp.department_id else None,
                "department_name": emp.department_id.name if emp.department_id else None,
                "job_id": emp.job_id.id if emp.job_id else None,
                "job_name": emp.job_id.name if emp.job_id else None,
                "status": "active" if emp.active else "inactive",
                "id_number": emp.id_number or None,
                "id_issued_date": str(emp.id_issued_date) if emp.id_issued_date else None,
                "id_issued_place": emp.id_issued_place or None,
                "permanent_address": emp.permanent_address or None,
                "temporary_address": emp.temporary_address or None,
                "tax_id": emp.tax_id or None,
                "insurance_id": emp.insurance_id or None,
                "bank_account": emp.bank_account or None,
                "contract": [
                    {
                        "id": c.id or None,
                        "name": c.name or None,
                        "contract_type": c.contract_type or None,
                        "contract_term": c.contract_term or None,
                        "date_start": str(c.date_start) if c.date_start else None,
                        "date_end": str(c.date_end) if c.date_end else None,
                        "wage": c.wage or None,
                        "bonus": c.bonus or None
                    }
                    for c in request.env['hr.contract.custom'].sudo().search([('employee_id', '=', emp.id)])],
            })

        return request.make_response(
        json.dumps({
            "data": data,
            "meta": {"page": page, "limit": limit, "total": total}
        }),
        headers=[('Content-Type', 'application/json')]
    )

    #get all employee code
    @http.route('/api/hr/employees/ids', type='http', auth='user', methods=['GET'], csrf=False)
    def get_all_employees(self, **kwargs):
        Employee = request.env['hr.employee'].sudo()
        employees = Employee.search([])
        data = [{"id": emp.id or None, "code": emp.code or None, "name": emp.name or None} for emp in employees]
        return request.make_response(
            json.dumps({"data": data}),
            headers=[('Content-Type', 'application/json')]
        )

    @http.route('/api/hr/employees/<int:employee_id>', type='http', auth='user', methods=['GET'], csrf=False)
    def get_employee(self, employee_id, **kwargs):
        Employee = request.env['hr.employee'].sudo()
        Contract = request.env['hr.contract.custom'].sudo()

        employee = Employee.search([('id', '=', employee_id)], limit=1)
        if not employee:
            return request.make_response(
                json.dumps({"error": "Employee not found"}),
                headers=[('Content-Type', 'application/json')],
                status=404
            )

        profile = {
            "id": employee.id or None,
            "code": employee.code or None,
            "name": employee.name or None,
            "birthday": str(employee.birthday) if employee.birthday else None,
            "gender": employee.gender or None,
            "work_phone": employee.work_phone or None,
            "work_email": employee.work_email or None,
            "department_id": employee.department_id.id if employee.department_id else None,
            "department_name": employee.department_id.name if employee.department_id else None,
            "job_id": employee.job_id.id if employee.job_id else None,
            "job_name": employee.job_id.name if employee.job_id else None,
            "id_number": employee.id_number or None,
            "id_issued_place": employee.id_issued_place or None,
            "id_issued_date": str(employee.id_issued_date) if employee.id_issued_date else None,
            "permanent_address": employee.permanent_address or None,
            "temporary_address": employee.temporary_address or None,
            "tax_id": employee.tax_id or None,
            "insurance_id": employee.insurance_id or None,
            "bank_account": employee.bank_account or None,
            "status": "active" if employee.active else "inactive",
            "created_at": str(employee.create_date) if employee.create_date else None,
            "updated_at": str(employee.write_date) if employee.write_date else None
        }

        contracts = []
        for c in Contract.search([('employee_id', '=', employee.id)]):
            contracts.append({
                "id": c.id or None,
                "name": c.name or None,
                "contract_type": c.contract_type or None,
                "contract_term": c.contract_term or None,
                "date_start": str(c.date_start) if c.date_start else None,
                "date_end": str(c.date_end) if c.date_end else None,
                "wage": c.wage or None,
                "bonus": c.bonus or None,
            })

        response = {
            "data": {
                "profile": profile,
                "contracts": contracts
            }   
        }

        return request.make_response(
            json.dumps(response),
            headers=[('Content-Type', 'application/json')]
        )

    @http.route('/api/hr/employees/<int:employee_id>', type='http', auth='user', methods=['PUT'], csrf=False, )
    def update_employee(self, employee_id, **kwargs):
        try:
            Employee = request.env['hr.employee'].sudo()
            employee = Employee.browse(employee_id)

            if not employee.exists():
                return request.make_json_response({"error": "Employee not found"}, status=404)

            # Lấy JSON body từ request
            try:
                data = request.httprequest.get_json(force=True, silent=True) or {}
            except Exception:
                data = {}

            allowed_fields = [
                'code', 'name', 'work_email', 'work_phone',
                'birthday', 'gender', 
                'department_id', 'job_id',
                'id_number', 'id_issued_place', 'id_issued_date',
                'permanent_address', 'temporary_address',
                'tax_id', 'insurance_id', 'bank_account',
                'contract',

            ]

            updates = {}
            updates_contract = {}
            ignored_fields = {}
            unchanged_fields = {}

            for field in allowed_fields:
                if field in data:
                    new_value = data[field]
                    old_value = None
                    if(field == 'contract'):
                        new_value = data[field][0] if isinstance(data[field], list) and len(data[field]) > 0 else None
                        old_value = request.env['hr.contract.custom'].sudo().search([('employee_id', '=', employee.id)], limit=1)
                    else:
                        old_value = employee[field]

                    if new_value in (None, "", False):
                        ignored_fields[field] = {
                            "reason": "empty",
                            "old_value": old_value,
                            "new_value": new_value
                        }
                        continue

                    if str(old_value) == str(new_value):
                        unchanged_fields[field] = {
                            "reason": "unchanged",
                            "old_value": old_value,
                            "new_value": new_value
                        }
                        continue
                    if(field == 'contract'):
                        contract_data = new_value
                        contract = old_value
                        print(contract_data)
                        if contract:
                            contract_updates = {}
                            for c_field in ['name', 'contract_type', 'contract_term', 'date_start', 'date_end', 'wage', 'bonus']:
                                if c_field in contract_data:
                                    c_new_value = contract_data[c_field]
                                    c_old_value = contract[c_field]
                                    if c_new_value not in (None, "", False) and str(c_old_value) != str(c_new_value):
                                        contract_updates[c_field] = c_new_value
                            if contract_updates:
                                updates_contract = contract_updates
                                contract.write(contract_updates)
                        else:
                            # Create new contract if none exists
                            request.env['hr.contract.custom'].sudo().create({
                                "employee_id": employee.id,
                                "name": contract_data.get("name"),
                                "contract_type": contract_data.get("contract_type"),
                                "contract_term": contract_data.get("contract_term"),
                                "date_start": contract_data.get("date_start"),
                                "date_end": contract_data.get("date_end"),
                                "wage": contract_data.get("wage"),
                                "bonus": contract_data.get("bonus")
                            })
                    else:
                        updates[field] = new_value

            if updates:
                employee.logs = (employee.logs or []) + [{
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "changes": {**updates, **updates_contract}
                }]
                employee.write(updates)
                return request.make_json_response({
                    "success": True,
                    "updated_fields": {**updates, **updates_contract},
                    "ignored_fields": ignored_fields,
                    "unchanged_fields": unchanged_fields
                }, status=200)
            else:
                return request.make_json_response({
                    "message": "No fields updated (empty or unchanged)",
                    "ignored_fields": ignored_fields,
                    "unchanged_fields": unchanged_fields,
                    "received_data": data
                }, status=200)

        except Exception as e:
            # Rollback transaction in case of error
            _logger.exception(f"Error updating employee {employee_id}: {str(e)}")
            request.env.cr.rollback()
            return request.make_json_response({"error": str(e)}, status=500)

    @http.route('/api/hr/employees/<int:employee_id>', type='http', auth='user', methods=['DELETE'], csrf=False, )
    def delete_employee(self, employee_id, **kwargs):
        Employee = request.env['hr.employee'].sudo()
        employee = Employee.search([('id', '=', employee_id)], limit=1)

        if not employee:
            return request.make_response(
                json.dumps({"error": "Employee not found"}),
                headers=[('Content-Type', 'application/json')],
                status=404
            )

        try:
            employee.write({'active': False})
        except Exception as e:
            return request.make_response(
                json.dumps({"error": str(e)}),
                headers=[('Content-Type', 'application/json')],
                status=500
            )

        return request.make_response(
            json.dumps({"message": "Employee marked inactive"}),
            headers=[('Content-Type', 'application/json')],
            status=200
        )

    @http.route('/api/hr/employees/export-template', type='http', auth='user', methods=['GET'], csrf=False)
    def download_import_template(self, **kwargs):
        try:
            Department = request.env['hr.department'].sudo()
            Job = request.env['hr.job'].sudo()
            Contract = request.env['hr.contract.custom'].sudo()

            departments = [d.name for d in Department.search([]) if d.name]
            jobs = [j.name for j in Job.search([]) if j.name]
            contract_types = list({c.contract_type for c in Contract.search([]) if c.contract_type})

            if not contract_types:
                contract_types = [
                    'Hợp đồng xác định thời hạn',
                    'Hợp đồng không xác định thời hạn',
                    'Hợp đồng thử việc'
                ]

            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})
            template_ws = workbook.add_worksheet('Template nhân viên')
            dept_ws = workbook.add_worksheet('Phòng ban')
            job_ws = workbook.add_worksheet('Chức vụ')
            contract_ws = workbook.add_worksheet('Loại hợp đồng')

            required_fmt = workbook.add_format({
                'bold': True, 'font_color': '#FF5722',
                'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True
            })
            optional_fmt = workbook.add_format({
                'bold': True,
                'border': 1, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True
            })

            headers = [
                ('Mã Nhân Viên', True),
                ('Tên Nhân Viên', True),
                ('Giới Tính', True),
                ('Ngày Sinh (YYYY-MM-DD)', True),
                ('Điện Thoại', True),
                ('Email', True),
                ('Ngày Vào Làm (YYYY-MM-DD)', True),
                ('Ngày Nghỉ Việc (YYYY-MM-DD)', False),
                ('Số CCCD', True),
                ('Ngày Cấp CCCD (YYYY-MM-DD)', True),
                ('Nơi Cấp CCCD', True),
                ('Địa Chỉ Thường Trú', True),
                ('Địa Chỉ Tạm Trú', False),
                ('Mã số thuế TNCN', False),
                ('Mã số BHXH', False),
                ('Tài Khoản Ngân Hàng', False),
                ('Phòng Ban', True),
                ('Chức Vụ', True),
                ('Loại Hợp Đồng', True),
                ('Thời Hạn Hợp Đồng', True),
                ('Mức Lương', False),
                ('Tiền Thưởng', False)
            ]

            for col, (title, req) in enumerate(headers):
                template_ws.set_column(col, col, 22)
                template_ws.write(0, col, title, required_fmt if req else optional_fmt)

            sample_row = [
                'MNV001', 'Nguyen Van A', 'Nam', '1990-05-15', '+84123456789', 'nguyenvana@company.com',
                '2025-01-01', '', '012345678901', '2015-03-10', 'Hà Nội',
                '123 Bạch Mai, Hai Bà Trưng, Hà Nội', '',
                '0123456789', '987654321', '1234567890123456 - Techcombank',
                (departments[0] if departments else ''),
                (jobs[0] if jobs else ''),
                (contract_types[0] if contract_types else ''),
                '12 tháng', '15000000', '2000000'
            ]
            for col, val in enumerate(sample_row):
                template_ws.write(1, col, val)

            # Ghi danh sách vào các sheet bên cạnh
            dept_ws.write(0, 0, 'Phòng ban')
            for i, name in enumerate(departments, start=1):
                dept_ws.write(i, 0, name)

            job_ws.write(0, 0, 'Chức vụ')
            for i, name in enumerate(jobs, start=1):
                job_ws.write(i, 0, name)

            contract_ws.write(0, 0, 'Loại hợp đồng')
            for i, name in enumerate(contract_types, start=1):
                contract_ws.write(i, 0, name)

            # Data validation cho các cột dropdown (áp dụng từ dòng 3 trở đi để không ghi đè mẫu)
            last_row = 1000
            if departments:
                template_ws.data_validation(2, 16, last_row, 16, {
                    'validate': 'list',
                    'source': f"='Phòng ban'!$A$2:$A${len(departments)+1}"
                })
            if jobs:
                template_ws.data_validation(2, 17, last_row, 17, {
                    'validate': 'list',
                    'source': f"='Chức vụ'!$A$2:$A${len(jobs)+1}"
                })
            if contract_types:
                template_ws.data_validation(2, 18, last_row, 18, {
                    'validate': 'list',
                    'source': f"='Loại hợp đồng'!$A$2:$A${len(contract_types)+1}"
                })

            workbook.close()
            output.seek(0)
            return request.make_response(
                output.read(),
                headers=[
                    ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                    ('Content-Disposition', 'attachment; filename=\"employee_import_template.xlsx\"')
                ]
            )
        except Exception as e:
            return request.make_response(
                json.dumps({'error': str(e)}),
                headers=[('Content-Type', 'application/json')],
                status=500
            )

    @http.route('/api/hr/employees/import', type='http', auth='user', methods=['POST'], csrf=False, )
    def import_employees_excel(self, **kwargs):
        try:
            # Get the uploaded file
            if not request.httprequest.files:
                return request.make_response(
                    json.dumps({"error": "No file uploaded"}),
                    headers=[('Content-Type', 'application/json')],
                    status=400
                )

            uploaded_file = request.httprequest.files.get('file')
            if not uploaded_file:
                return request.make_response(
                    json.dumps({"error": "No file found in request"}),
                    headers=[('Content-Type', 'application/json')],
                    status=400
                )

            # Validate file type
            if not uploaded_file.filename.endswith(('.xlsx', '.xls')):
                return request.make_response(
                    json.dumps({"error": "Invalid file type. Only .xlsx and .xls files are allowed"}),
                    headers=[('Content-Type', 'application/json')],
                    status=400
                )

            # Read the Excel file
            file_content = uploaded_file.read()
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            worksheet = workbook.active

            # Expected columns mapping
            expected_columns = {
                'A': 'Mã Nhân Viên',
                'B': 'Tên Nhân Viên', 
                'C': 'Giới Tính',
                'D': 'Ngày Sinh (YYYY-MM-DD)',
                'E': 'Điện Thoại',
                'F': 'Email',
                'G': 'Ngày Vào Làm (YYYY-MM-DD)',
                'H': 'Ngày Nghỉ Việc (YYYY-MM-DD)',
                'I': 'Số CCCD',
                'J': 'Ngày Cấp CCCD (YYYY-MM-DD)',
                'K': 'Nơi Cấp CCCD',
                'L': 'Địa Chỉ Thường Trú',
                'M': 'Địa Chỉ Tạm Trú',
                'N': 'Mã Số Thuế TNCN',
                'O': 'Mã Số BHXH',
                'P': 'Tài Khoản Ngân Hàng',
                'Q': 'Phòng Ban',
                'R': 'Chức Vụ',
                'S': 'Loại Hợp Đồng',
                'T': 'Thời Hạn Hợp Đồng',
                'U': 'Mức Lương',
                'V': 'Tiền Thưởng'
            }

            results = {
                'success': [],
                'errors': [],
                'total_processed': 0,
                'total_created': 0,
                'total_errors': 0
            }

            # Process each row (skip header row)
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    results['total_processed'] += 1
                    
                    # Extract data from row
                    employee_data = {}
                    for col_idx, (col_letter, field_name) in enumerate(expected_columns.items()):
                        if col_idx < len(row):
                            value = row[col_idx]
                            if value is not None:
                                if field_name in ['birthday', 'id_issued_date', 'start_date', 'end_date']:
                                    # Handle date fields
                                    if isinstance(value, datetime):
                                        employee_data[field_name] = value.strftime('%Y-%m-%d')
                                    elif isinstance(value, str):
                                        employee_data[field_name] = value.strip()
                                else:
                                    employee_data[field_name] = str(value).strip() if value else None

                    # Validate required fields
                    required_fields = ['Mã Nhân Viên', 'Tên Nhân Viên', 'Ngày Sinh (YYYY-MM-DD)', 'Giới Tính', 'Điện Thoại', 'Email',
                                       'Phòng Ban', 'Chức Vụ', 'Số CCCD', 'Nơi Cấp CCCD',
                                       'Ngày Cấp CCCD (YYYY-MM-DD)', 'Địa Chỉ Thường Trú', 'Loại Hợp Đồng', 'Thời Hạn Hợp Đồng']

                    missing_fields = []
                    for field in required_fields:
                        if not employee_data.get(field):
                            missing_fields.append(field)

                    if missing_fields:
                        results['errors'].append({
                            'row': row_num,
                            'error': f"Missing required fields: {', '.join(missing_fields)}",
                            'data': employee_data
                        })
                        results['total_errors'] += 1
                        continue

                    # Find department and job by name
                    department = request.env['hr.department'].sudo().search([
                        ('name', '=', employee_data['Phòng Ban'])
                    ], limit=1)
                    
                    if not department:
                        results['errors'].append({
                            'row': row_num,
                            'error': f"Department '{employee_data['department_name']}' not found",
                            'data': employee_data
                        })
                        results['total_errors'] += 1
                        continue

                    job = request.env['hr.job'].sudo().search([
                        ('name', '=', employee_data['Chức Vụ'])
                    ], limit=1)
                    
                    if not job:
                        results['errors'].append({
                            'row': row_num,
                            'error': f"Job '{employee_data['job_name']}' not found",
                            'data': employee_data
                        })
                        results['total_errors'] += 1
                        continue

                    # Check if employee with same email or id_number already exists
                    existing_employee = request.env['hr.employee'].sudo().search([
                        '|',
                        ('code', '=', employee_data['Mã Nhân Viên']),
                        ('id_number', '=', employee_data['Số CCCD']),
                    ], limit=1)

                    if existing_employee:
                        results['errors'].append({
                            'row': row_num,
                            'error': f"Mã Nhân Viên '{employee_data['Mã Nhân Viên']}' or Số CCCD '{employee_data['Số CCCD']}' already exists",
                            'data': employee_data
                        })
                        results['total_errors'] += 1
                        continue

                    # Create employee
                    employee = request.env['hr.employee'].sudo().create({
                        'code': employee_data['Mã Nhân Viên'],
                        'name': employee_data['Tên Nhân Viên'],
                        'birthday': employee_data['Ngày Sinh (YYYY-MM-DD)'],
                        'gender': employee_data['Giới Tính'],
                        'work_phone': employee_data['Điện Thoại'],
                        'work_email': employee_data['Email'],
                        'department_id': department.id,
                        'job_id': job.id,
                        'id_number': employee_data['Số CCCD'],
                        'id_issued_place': employee_data['Nơi Cấp CCCD'],
                        'id_issued_date': employee_data['Ngày Cấp CCCD (YYYY-MM-DD)'],
                        'permanent_address': employee_data['Địa Chỉ Thường Trú'],
                        'temporary_address': employee_data.get('Địa Chỉ Tạm Trú'),
                        'tax_id': employee_data.get('Mã số thuế TNCN'),
                        'insurance_id': employee_data.get('Mã số BHXH'),
                        'bank_account': employee_data.get('Tài Khoản Ngân Hàng'),
                        'active': True
                    })

                    # Create contract type if not exists

                    contract_type_record = request.env['hr.contract.custom'].sudo().create({
                        'name': employee_data['Loại Hợp Đồng'],
                        'contract_type': employee_data['Loại Hợp Đồng'],
                        'contract_term': employee_data.get('Thời Hạn Hợp Đồng') or None,
                        'employee_id': employee.id,
                        'date_start': employee_data.get('Ngày Vào Làm (YYYY-MM-DD)'),
                        'date_end': employee_data.get('Ngày Nghỉ Việc (YYYY-MM-DD)') or None,
                        'wage': employee_data.get('Mức Lương') or 0,
                        'bonus': employee_data.get('Tiền Thưởng') or 0,
                    })

                    results['success'].append({
                        'row': row_num,
                        'employee_id': employee.id,
                        'name': employee.name,
                        'email': employee.work_email
                    })
                    results['total_created'] += 1

                except Exception as e:
                    _logger.exception(f"Error processing row {row_num}: {str(e)}")
                    results['errors'].append({
                        'row': row_num,
                        'error': str(e),
                        'data': employee_data if 'employee_data' in locals() else {}
                    })
                    results['total_errors'] += 1

            return request.make_response(
                json.dumps(results, ensure_ascii=False),
                headers=[('Content-Type', 'application/json')]
            )

        except Exception as e:
            # Rollback transaction in case of error
            request.env.cr.rollback()
            return request.make_response(
                json.dumps({"error": f"Import failed: {str(e)}"}, ensure_ascii=False),
                headers=[('Content-Type', 'application/json')],
                status=500
            )

    @http.route('/api/hr/employees/export', type='http', auth='user', methods=['GET'], csrf=False)
    def export_employees(self, **kwargs):
        try:
            q = kwargs.get('q', '')
            department_id = kwargs.get('department_id')
            job_id = kwargs.get('job_id')
            gender = kwargs.get('gender', '').lower()

            domain = []
            if q:
                domain.append(('name', 'ilike', q))
            if department_id:
                domain.append(('department_id', '=', int(department_id)))
            if job_id:
                domain.append(('job_id', '=', int(job_id)))
            if gender:
                domain.append(('gender', '=', gender))

            Employee = request.env['hr.employee'].sudo()
            Contract = request.env['hr.contract.custom'].sudo()
            employees = Employee.search(domain, order='code asc')

            # Collect contract data in batch
            contracts = Contract.search([('employee_id', 'in', employees.ids)])
            contracts_by_emp = {}
            for c in contracts:
                contracts_by_emp.setdefault(c.employee_id.id, []).append(c)

            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})

            # Sheet 1: Employees
            emp_ws = workbook.add_worksheet('Employees')
            header_fmt = workbook.add_format({
                'bold': True, 'bg_color': '#1E88E5', 'font_color': 'white',
                'border': 1, 'align': 'center', 'valign': 'vcenter'
            })
            text_wrap = workbook.add_format({'text_wrap': True, 'valign': 'top'})

            headers = [
                'Mã Nhân Viên',
                'Tên Nhân Viên', 
                'Giới Tính',
                'Ngày Sinh (YYYY-MM-DD)',
                'Điện Thoại',
                'Email',
                'Ngày Vào Làm (YYYY-MM-DD)',
                'Ngày Nghỉ Việc (YYYY-MM-DD)',
                'Số CCCD',
                'Ngày Cấp CCCD (YYYY-MM-DD)',
                'Nơi Cấp CCCD',
                'Địa Chỉ Thường Trú',
                'Địa Chỉ Tạm Trú',
                'Mã Số Thuế TNCN',
                'Mã Số BHXH',
                'Tài Khoản Ngân Hàng',
                'Phòng Ban',
                'Chức Vụ',
                'Loại Hợp Đồng',
                'Thời Hạn Hợp Đồng',
                'Mức Lương',
                'Tiền Thưởng'
            ]
            for col, h in enumerate(headers):
                emp_ws.write(0, col, h, header_fmt)
                emp_ws.set_column(col, col, 20)

            for row_idx, emp in enumerate(employees, start=1):
                emp_ws.write(row_idx, 0, emp.code or '')
                emp_ws.write(row_idx, 1, emp.name or '')
                emp_ws.write(row_idx, 2, emp.gender or '')
                emp_ws.write(row_idx, 3, emp.birthday and emp.birthday.strftime('%Y-%m-%d') or '')
                emp_ws.write(row_idx, 4, emp.work_phone or '')
                emp_ws.write(row_idx, 5, emp.work_email or '')
                #get contract start and end date
                contract = contracts_by_emp.get(emp.id, [])

                emp_ws.write(row_idx, 6, (contract and contract[0].date_start.strftime('%Y-%m-%d')) or '')
                emp_ws.write(row_idx, 7, ((contract and contract[0].date_end) or '').strftime('%Y-%m-%d') if (contract and contract[0].date_end) else '')
                emp_ws.write(row_idx, 8, emp.id_number or '')
                emp_ws.write(row_idx, 9, str(emp.id_issued_date or '') or '')
                emp_ws.write(row_idx, 10, emp.id_issued_place or '')
                emp_ws.write(row_idx, 11, emp.permanent_address or '', text_wrap)
                emp_ws.write(row_idx, 12, emp.temporary_address or '', text_wrap)
                emp_ws.write(row_idx, 13, emp.tax_id or '')
                emp_ws.write(row_idx, 14, emp.insurance_id or '')
                emp_ws.write(row_idx, 15, emp.bank_account or '')
                emp_ws.write(row_idx, 16, emp.department_id.name or '')
                emp_ws.write(row_idx, 17, emp.job_id.name or '')
                emp_ws.write(row_idx, 18, contract and contract[0].contract_type or '')
                emp_ws.write(row_idx, 19, contract and contract[0].contract_term or '')
                emp_ws.write(row_idx, 20, contract and contract[0].wage or 0)
                emp_ws.write(row_idx, 21, contract and contract[0].bonus or 0)

            workbook.close()
            output.seek(0)

            filename = f"employees_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return request.make_response(
                output.read(),
                headers=[
                    ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                    ('Content-Disposition', f'attachment; filename="{filename}"')
                ]
            )
        except Exception as e:
            _logger.exception("Employee export failed")
            return request.make_response(
                json.dumps({'error': str(e)}),
                headers=[('Content-Type', 'application/json')],
                status=500
            )

    #delete batch employee by ids
    @http.route('/api/hr/employees/batch-delete', type='http', auth='user', methods=['POST'], csrf=False, )
    def batch_delete_employees(self, **kwargs):
        try:
            # Lấy JSON body từ request
            try:
                data = request.httprequest.get_json(force=True, silent=True) or {}
            except Exception:
                data = {}

            employee_ids = data.get('employee_ids', [])
            if not employee_ids or not isinstance(employee_ids, list):
                return request.make_response(
                    json.dumps({"error": "employee_ids is required and must be a list of IDs"}),
                    headers=[('Content-Type', 'application/json')],
                    status=400
                )

            Employee = request.env['hr.employee'].sudo()
            employees = Employee.search([('id', 'in', employee_ids)])

            if not employees:
                return request.make_response(
                    json.dumps({"error": "No matching employees found for the provided IDs"}),
                    headers=[('Content-Type', 'application/json')],
                    status=404
                )

            # Mark employees as inactive
            employees.write({'active': False})

            return request.make_response(
                json.dumps({
                    "message": f"Marked {len(employees)} employees as inactive successfully",
                    "employee_ids": employees.ids
                }),
                headers=[('Content-Type', 'application/json')],
                status=200
            )
        except Exception as e:
            # Rollback transaction in case of error
            request.env.cr.rollback()
            return request.make_response(
                json.dumps({"error": str(e)}),
                headers=[('Content-Type', 'application/json')],
                status=500
            )
        
